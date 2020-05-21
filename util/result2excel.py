from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font


def result2excel(project, apiName, CaseName, msg):
    global result_index
    project = project
    case_excel = Path(f'projects/{project}/data/{project}.xlsx')
    # case_excel = Path(f'../projects/{project}/data/{project}.xlsx')

    wb = load_workbook(case_excel)
    ws = wb[apiName]

    case_name_list = list(ws.iter_cols(values_only=True))
    case_name_tuple = case_name_list[0]
    case_name_index = case_name_tuple.index(CaseName)  # 行坐标
    max_column = ws.max_column
    for result_index in range(max_column):
        if case_name_list[result_index][1] == 'result':  # 列坐标
            break
    ws.cell(row=case_name_index + 1, column=result_index + 1, value=msg)
    if msg =='pass':
        ws.cell(row=case_name_index + 1, column=result_index + 1).font = Font(color='32CD32')
    else:
        ws.cell(row=case_name_index + 1, column=result_index + 1).font = Font(color='FF0000')
    wb.save(case_excel)
    wb.close()

if __name__ == '__main__':
    result2excel('BigData', 'phone_check', '全部入参正常-问问：单独', 'pass1')
