import yaml
from pathlib import Path
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

def get_yaml(yaml_file):
    with open(yaml_file, encoding='UTF-8') as yf:
        api_info = yaml.load(yf)
        # pprint(api_info)
        return api_info

def param_to_excel(api_info, case_excel):
    wb = load_workbook(case_excel)
    sheet_name_list = wb.get_sheet_names()
    if not api_info['apiName'] in sheet_name_list:
        print(f'创建新的sheet:{api_info["apiName"]}')
        wb.create_sheet(api_info['apiName'], index=0)
    else:
        print(f'已存在的sheet：{api_info["apiName"]}')
        return
    sheet = wb[api_info['apiName']]
    # 入参个数
    # request_params_len = len(api_info['parmas'])
    # 断言个数
    response_assert_len = len(api_info['response_assert'])

    Y_params_list = []  # 必填入参列表
    N_params_list = []  # 选填入参列表
    for k, v in api_info['parmas'].items():
        if v['Required'] in ['Y', 'y', 'yes', 'Yes', 'YES']:
            Y_params_list.append(k)
        else:
            N_params_list.append(k)

    sheet['B1'] = '必填入参'
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=2 + len(Y_params_list) - 1)
    sheet['B1'].fill = PatternFill('solid', fgColor='FF3030')
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list), value='选填入参')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list), end_row=1,
                      end_column=2 + len(Y_params_list) + len(N_params_list) - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list)).fill = PatternFill('solid', fgColor='8FBC8F')
    sheet.cell(row=1, column=2 + len(Y_params_list)).alignment = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list), value='响应断言')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list) + len(N_params_list), end_row=1,
                      end_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list)).fill = PatternFill('solid', fgColor='CDAD00')
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list)).alignment = Alignment(horizontal='center',
                                                                                                vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len, value='测试结果')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len,
                      end_row=1, end_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len + 2 - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len).fill = PatternFill(
        'solid', fgColor='98FB98')
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len).alignment = Alignment(
        horizontal='center', vertical='center')

    title_list = ['CaseName'] + Y_params_list + N_params_list + api_info['response_assert'] + ['exec', 'result']
    sheet.append(title_list)
    # sheet.row_dimensions[2].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(case_excel)
    wb.close()

def main(project_name, case_name=None):
    yaml_conf_folder = Path(f'../projects/{project_name}/conf/')
    case_excel_path =  Path(f'../projects/{project_name}/data/{project_name}.xlsx')
    if case_name:
        yaml_list = [yaml_conf_folder/(case_name+'.yaml')]
    else:
        yaml_list = list(yaml_conf_folder.glob('*.yaml'))
    for yaml_file in yaml_list:
        api_info = get_yaml(yaml_file)
        param_to_excel(api_info,case_excel_path)
    print('Done!!!')

if __name__ == '__main__':
    # main('BigData','phone_check') #给了项目名和具体案例名，生成具体案例的excel sheet模板
    main('BigData') #只给项目名，找到项目的conf目录，为所有yaml生成excel sheet,已生成的自动跳过