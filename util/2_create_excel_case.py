import yaml
from pathlib import Path
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


def get_yaml(yaml_file):
    with open(yaml_file, encoding='UTF-8') as yf:
        api_info = yaml.load(yf)
        # pprint(api_info)
        return api_info


def param_to_excel(api_info, case_excel):
    wb = load_workbook(case_excel)
    sheet_name_list = wb.get_sheet_names()
    if not api_info['apiName'] in sheet_name_list:
        print(f'创建新的sheet:{api_info["apiName"]}')
        wb.create_sheet(api_info['apiName'], index=0)
    else:
        print(f'已存在的sheet：{api_info["apiName"]}')
        return
    sheet = wb[api_info['apiName']]
    # 入参个数
    # request_params_len = len(api_info['parmas'])
    # 断言个数
    response_assert_len = len(api_info['response_assert'])

    Y_params_list = []  # 必填入参列表
    N_params_list = []  # 选填入参列表
    for k, v in api_info['parmas'].items():
        if v['Required'] in ['Y', 'y', 'yes', 'Yes', 'YES']:
            Y_params_list.append(k)
        else:
            N_params_list.append(k)

    sheet['B1'] = '必填入参'
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=2 + len(Y_params_list) - 1)
    sheet['B1'].fill = PatternFill('solid', fgColor='FF3030')
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list), value='选填入参')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list), end_row=1,
                      end_column=2 + len(Y_params_list) + len(N_params_list) - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list)).fill = PatternFill('solid', fgColor='8FBC8F')
    sheet.cell(row=1, column=2 + len(Y_params_list)).alignment = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list), value='响应断言')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list) + len(N_params_list), end_row=1,
                      end_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list)).fill = PatternFill('solid', fgColor='CDAD00')
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list)).alignment = Alignment(horizontal='center',
                                                                                                vertical='center')

    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len, value='测试结果')
    sheet.merge_cells(start_row=1, start_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len,
                      end_row=1, end_column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len + 2 - 1)
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len).fill = PatternFill(
        'solid', fgColor='98FB98')
    sheet.cell(row=1, column=2 + len(Y_params_list) + len(N_params_list) + response_assert_len).alignment = Alignment(
        horizontal='center', vertical='center')

    title_list = ['CaseName'] + Y_params_list + N_params_list + api_info['response_assert'] + ['exec', 'result']
    sheet.append(title_list)
    # sheet.row_dimensions[2].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(case_excel)
    wb.close()


def create_pre_test(project_name, api_info):
    apiName = api_info['apiName']
    pre_test_path = Path(f'../projects/{project_name}/pre_test/')
    pre_test_list = [pj.name for pj in pre_test_path.glob('*') if pj.is_file() and pj.name != '__init__.py']
    if f'{apiName}.py' in pre_test_list:
        print(f'该pre_test已存在： {apiName}')
    else:
        print(f'生成pre_test:{apiName}')
        with open(Path(f'../projects/{project_name}/pre_test/{apiName}.py'), mode='w', encoding='UTF-8') as f:
            pre_code = '''
def handdle_data(data):



    return data
    
'''
            f.write(pre_code)


def create_case(project_name, api_info):
    apiName = api_info['apiName']
    assert_param_list = api_info['response_assert']
    assert_content = ''
    for assert_param in assert_param_list:
        assert_content += f"assert responce_assert['{assert_param}'] == res['{assert_param}']\n                "

    case_path = Path(f'../projects/{project_name}/test_cases')
    case_list = [pj.name for pj in case_path.glob('*') if pj.is_file() and pj.name != '__init__.py']
    if f'test_{apiName}.py' in case_list:
        print(f'该案例已存在：{apiName}')
    else:
        print(f'生成py案例：{apiName}')

        with open(case_path / f'test_{apiName}.py', mode='w', encoding='UTF-8') as f:
            case_content = f'''
import pytest
import requests

from util.result2excel import result2excel


class Test{apiName}:
    apiName = '{apiName}'

    # scope='module'意味着该fixture只在案例开始前运行一次
    @pytest.fixture(scope='module', autouse=True)
    def cls_handdler(self):
        print('pre do once')

    # 类级别的事务fixture被标记为autouse=true，这意味着类中的所有测试方法将使用这个fixture，而不需要在测试函数签名中声明它
    @pytest.fixture(autouse=True)
    def handdler(self):
        print('pre do')
        yield
        print('post do')

    # @pytest.mark.skip(reason="no way of currently testing this")a
    # @pytest.mark.skipif(3<2, reason="3>2")
    def test_api(self, param):
        print(param)
        if param['request_info']['apiName'] != Test{apiName}.apiName:
            print('apiName不匹配！')
        else:
            url = param['request_info']['request_url']
            header = param['request_info']['headers']
            data = param['request_params']
            responce_assert = param['response_assert']
            try:
                res = requests.post(url, json=data, headers=header).json()
                {assert_content}
                result2excel(__class__.__module__.split('.')[1], __class__.apiName, param['CaseName'], 'pass')
            except AssertionError as e:
                result2excel(__class__.__module__.split('.')[1], __class__.apiName, param['CaseName'], str(e))
                raise AssertionError(e)

'''
            f.write(case_content)


def main(project_name, case_name=None):
    yaml_conf_folder = Path(f'../projects/{project_name}/conf/')
    case_excel_path = Path(f'../projects/{project_name}/data/{project_name}.xlsx')
    if case_name:
        yaml_list = [yaml_conf_folder / (case_name + '.yaml')]
    else:
        yaml_list = list(yaml_conf_folder.glob('*.yaml'))
    for yaml_file in yaml_list:
        print(yaml_file)
        api_info = get_yaml(yaml_file)
        param_to_excel(api_info, case_excel_path)  # yam数据生成excel案例
        create_pre_test(project_name, api_info)  # yaml信息创建pre_test
        create_case(project_name, api_info)  # yaml 信息创建test_case
    print('Done!!!')


if __name__ == '__main__':
    '''
    前提是已经使用1_create_project.py生成的项目目录
    给了项目名和具体案例名，生成具体案例的excel sheet模板、pre_test、testcase，已生成的自动跳过
    只给项目名，找到项目的conf目录，为所有yaml生成excel sheet、pre_test、testcase，已生成的自动跳过
    '''
    # main('BigData','phone_check') #给了项目名和具体案例名
    main('VBA')  # 只给项目名
